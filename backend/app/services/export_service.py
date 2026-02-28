import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_excel(tasks: list[dict], custom_columns: list[dict] | None = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = [
        "ID",
        "Task Name",
        "Description",
        "Owner",
        "Email",
        "Start Date",
        "Due Date",
        "Status",
        "Priority",
    ]

    # Append custom column headers
    if custom_columns:
        for col in custom_columns:
            headers.append(col["display_name"])

    headers.append("Created At")

    header_fill = PatternFill(start_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, task in enumerate(tasks, 2):
        col_idx = 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("id")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("task_name")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("description")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("owner")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("email")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=str(task.get("start_date") or "")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=str(task.get("due_date") or "")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("status")); col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=task.get("priority")); col_idx += 1

        # Custom field values
        if custom_columns:
            cf = json.loads(task.get("custom_fields") or "{}")
            for custom_col in custom_columns:
                value = cf.get(custom_col["field_key"], "")
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                col_idx += 1

        ws.cell(row=row_idx, column=col_idx, value=str(task.get("created_at") or ""))

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
